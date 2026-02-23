from __future__ import annotations

import os
from dataclasses import asdict
from datetime import datetime
from typing import Dict, Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from ..reconciliation.models import ReconcileRow, ReconcileStatus


# Simple fills (keep readable on default Excel themes)
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GRAY = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

HEADER_FONT = Font(bold=True)


def _fill_for_status(status: ReconcileStatus) -> PatternFill:
    if status == ReconcileStatus.MATCHED:
        return FILL_GREEN
    if status in (ReconcileStatus.AMOUNT_MISMATCH, ReconcileStatus.TIME_MISMATCH):
        return FILL_YELLOW
    if status in (ReconcileStatus.MISSING_IN_INTERNAL, ReconcileStatus.MISSING_IN_STRIPE):
        return FILL_RED
    return FILL_GRAY


def write_report(rows: List[ReconcileRow], summary: Dict[str, int], out_path: str) -> str:
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)

    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "summary"
    ws_summary["A1"] = "metric"
    ws_summary["B1"] = "value"
    ws_summary["A1"].font = HEADER_FONT
    ws_summary["B1"].font = HEADER_FONT

    r = 2
    for k in sorted(summary.keys()):
        ws_summary[f"A{r}"] = k
        ws_summary[f"B{r}"] = summary[k]
        r += 1

    ws_summary[f"A{r+1}"] = "generated_at_utc"
    ws_summary[f"B{r+1}"] = datetime.utcnow().isoformat() + "Z"

    # Details sheet
    ws = wb.create_sheet("reconciliation")
    headers = [
        "transaction_id",
        "status",
        "confidence",
        "currency",
        "stripe_amount",
        "internal_amount",
        "stripe_time",
        "internal_time",
        "notes",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT

    # Deterministic ordering
    rows_sorted = sorted(rows, key=lambda x: x.transaction_id)

    for row in rows_sorted:
        ws.append([
            row.transaction_id,
            row.status.value,
            float(row.confidence),
            row.currency,
            row.stripe_amount,
            row.internal_amount,
            row.stripe_time.isoformat() if row.stripe_time else None,
            row.internal_time.isoformat() if row.internal_time else None,
            row.notes,
        ])

    # Apply row fills
    for i in range(2, 2 + len(rows_sorted)):
        status_val = ws.cell(row=i, column=2).value
        fill = _fill_for_status(ReconcileStatus(status_val))
        for j in range(1, len(headers) + 1):
            ws.cell(row=i, column=j).fill = fill

    # Autosize columns (basic)
    for col in range(1, len(headers) + 1):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[get_column_letter(col)].width = min(60, max(12, max_len + 2))

    wb.save(out_path)
    return out_path
